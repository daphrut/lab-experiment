# Function to clean affected variables:
#   (1) Auto-discover all *_aff_vars columns in the main dataset
#   (2) Build a long dataframe: labgroupid × trigger_var × affected_var × original_value_clean
#   (3) Append new cases (status "Unchecked") to the specified Excel sheet
#   (4) Pull back cleaned values where value_changed == "Y" into *_clean columns

import pandas as pd
from openpyxl import load_workbook
import os

def clean_affected_vars(df, file_name, sheet_name, data_dict=None, id_cols=None):
    """
    Export affected-variable cases to an Excel review sheet and pull back
    any already-cleaned values into the main dataframe.

    Parameters
    ----------
    df : pd.DataFrame
        Main dataset. Must contain all `id_cols` and any number of
        *_aff_vars columns produced by clean_unique_values.
    file_name : path-like
        Path to the affected_variables_cleaning Excel workbook.
    sheet_name : str
        Sheet name within the workbook to work with.
    data_dict : pd.DataFrame, optional
        Data dictionary with "Variable" and "No variables" columns, used to
        expand multi-instance fields (e.g. source -> source_1, source_2, …).
    id_cols : list[str], optional
        Columns that uniquely identify an observation. Defaults to
        ["labgroupid"]. For panel data, pass for example:
        ["labgroupid", "equipment", "type_no", "survey"].

    Returns
    -------
    pd.DataFrame
        Updated df with *_clean columns overwritten for matching observations where
        value_changed == "Y" in the cleaning workbook.
    """
    
    # Input validation for id_cols
    if id_cols is None:
        id_cols = ["labgroupid"]

    missing_id_cols = [c for c in id_cols if c not in df.columns]
    if missing_id_cols:
        raise ValueError(
            f"Missing identifier column(s) in df: {missing_id_cols}. "
            f"Expected id_cols={id_cols}."
        )

    if df[id_cols].isna().any(axis=None):
        na_counts = df[id_cols].isna().sum()
        na_counts = na_counts[na_counts > 0].to_dict()
        raise ValueError(
            "Identifier columns contain missing values, cannot safely match rows. "
            f"NA counts by id column: {na_counts}"
        )

    if df.duplicated(subset=id_cols).any():
        dup_count = int(df.duplicated(subset=id_cols).sum())
        raise ValueError(
            f"Identifier columns are not unique ({dup_count} duplicate row(s)). "
            f"Use id_cols that uniquely identify observations: {id_cols}"
        )

    df = df.copy()

    # Normalize legacy blank strings in *_aff_vars columns to true missing values.
    aff_var_cols = [c for c in df.columns if c.endswith("_aff_vars")]
    for col in aff_var_cols:
        df[col] = df[col].astype("string").str.strip()
        df.loc[df[col] == "", col] = pd.NA

    # Step 1: Collect rows where any *_aff_vars column is non-empty
    parts = []
    for aff_var_col in aff_var_cols:
        trigger_var = aff_var_col.replace("_aff_vars", "")
        subset = df.loc[
            df[aff_var_col].notna(),
            id_cols + [aff_var_col]
        ].copy()
        if subset.empty:
            continue
        subset = subset.rename(columns={aff_var_col: "aff_vars"})
        subset["trigger_var"] = trigger_var
        parts.append(subset)

    if not parts:
        print("No affected variable cases found.")
        return df

    df_aff = pd.concat(parts, ignore_index=True)

    # Step 2: Explode comma-separated affected var names into one row each
    df_aff["affected_var"] = df_aff["aff_vars"].str.split(",")
    df_aff = df_aff.explode("affected_var").reset_index(drop=True)
    df_aff["affected_var"] = df_aff["affected_var"].str.strip()
    df_aff = df_aff.drop(columns="aff_vars")

    # Step 3 (optional): Expand multi-instance fields using data_dict
    if data_dict is not None:
        df_aff = df_aff.merge(
            data_dict[["Variable", "No variables"]],
            left_on="affected_var", right_on="Variable", how="left"
        )
        df_aff["No variables"] = df_aff["No variables"].fillna(1).astype(int)
        df_aff["affected_var"] = df_aff.apply(
            lambda r: [f"{r['affected_var']}_{i}" for i in range(1, r["No variables"] + 1)]
            if r["No variables"] > 1 else r["affected_var"],
            axis=1
        )
        df_aff = df_aff.explode("affected_var").reset_index(drop=True)
        df_aff = df_aff.drop(columns=["Variable", "No variables"], errors="ignore")

    # Step 4: Attach the current cleaned value of each affected variable if exists
    def _extract_original_value(row):
        var = row["affected_var"]
        clean_var = f"{var}_clean"

        # Prefer the *_clean column when present so reviews/editing always target cleaned values.
        value_col = clean_var if clean_var in df.columns else var
        if value_col not in df.columns:
            return pd.NA
        mask = pd.Series(True, index=df.index)
        for c in id_cols:
            mask &= df[c] == row[c]
        matches = df.loc[mask, value_col]
        if matches.empty:
            return pd.NA
        return matches.iloc[0]

    df_aff["original_value"] = df_aff.apply(_extract_original_value, axis=1)

    # Step 5: Build new-cases frame with standard columns
    new_cases = df_aff[id_cols + ["trigger_var", "affected_var", "original_value"]].copy()
    new_cases["cleaned_value"] = ""
    new_cases["check_status"] = "Unchecked"
    new_cases["value_changed"] = ""

    # Step 6: Load existing sheet from workbook
    if not os.path.exists(file_name):
        raise FileNotFoundError(f"Workbook not found: {file_name}. Run create_empty_aff_vars_sheet first.")
    wb = load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Run create_empty_aff_vars_sheet first.")
    ws = wb[sheet_name]

    rows = list(ws.values)
    if rows:
        sheet_cols = list(rows[0])
    else:
        sheet_cols = []

    if len(rows) > 1:
        existing = pd.DataFrame(rows[1:], columns=sheet_cols)
    else:
        existing = pd.DataFrame(columns=sheet_cols)

    required_sheet_cols = id_cols + [
        "trigger_var", "affected_var", "original_value",
        "cleaned_value", "check_status", "value_changed"
    ]
    missing_sheet_cols = [c for c in required_sheet_cols if c not in existing.columns]
    if missing_sheet_cols:
        raise ValueError(
            f"Sheet '{sheet_name}' is missing required column(s): {missing_sheet_cols}. "
            "Create/update the sheet with create_empty_aff_vars_sheet using matching id_cols."
        )

    # Step 7: Identify new rows and append them to the sheet
    key_cols = id_cols + ["trigger_var", "affected_var"]
    if not existing.empty:
        to_append = new_cases.merge(existing[key_cols], on=key_cols, how="left", indicator=True)
        to_append = to_append[to_append["_merge"] == "left_only"].drop(columns="_merge")
    else:
        to_append = new_cases.copy()

    if not to_append.empty:
        col_order = list(existing.columns)
        rows_to_write = to_append.copy()
        for c in col_order:
            if c not in rows_to_write.columns:
                rows_to_write[c] = ""
        rows_to_write = rows_to_write[col_order]
        rows_to_write = rows_to_write.where(rows_to_write.notna(), None)

        next_row = ws.max_row + 1
        for _, row in rows_to_write.iterrows():
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=next_row, column=col_idx, value=val)
            next_row += 1
        wb.save(file_name)
        print(f"Appended {len(to_append)} new affected variable case(s) to '{sheet_name}'.")
    else:
        print("No new affected variable cases found.")
    wb.close()

    # Step 8: Pull back cleaned values only if value_changed == "Y"
    if not existing.empty and "value_changed" in existing.columns:
        changed = existing[
            existing["value_changed"].astype(str).str.strip().str.upper() == "Y"
        ]
        for _, row in changed.iterrows():
            clean_col = f"{row['affected_var']}_clean"
            if clean_col not in df.columns:
                src_col = row["affected_var"]
                if src_col in df.columns:
                    df[clean_col] = df[src_col]
                else:
                    df[clean_col] = pd.NA

            mask = pd.Series(True, index=df.index)
            for c in id_cols:
                mask &= df[c] == row[c]
            df.loc[mask, clean_col] = row["cleaned_value"]
    
    # Step 9: Print report of number of changed values
    num_changed = existing[
        existing["value_changed"].astype(str).str.strip().str.upper() == "Y"
    ].shape[0]
    print(f"Pulled back cleaned values for {num_changed} affected variable case(s) where value_changed == 'Y'.")

    return df