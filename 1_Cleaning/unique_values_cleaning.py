### TO DO: deal with multiple variables and one sheet (e.g. _1, _2, etc.)


# Function to clean unique values:
#       (1) Merge with the cleaning worksheet to check if value-comment(-free text-free text comment) 
#           combination has already been added to the cleaning worksheet
#       (2) If not, add to the bottom of the cleaning worksheet with status "Unchecked" for review
#       (3) If cleaned, pull cleaned value and status for use in cleaning the main dataset
#       (4) If specified, produce short report of cleaning progress (e.g. no unique values, % cleaned, 
#           % pending, % exclude, etc.)

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def clean_unique_values(df, file_name, var_name, sheet_name, 
                        dtype="string",
                        comment=False, 
                        free_text=False, 
                        mc_fc_vars=False,
                        report=False):
    
    # Step 1: Prepare the data for merging with the cleaning sheet

    # List of cols to merge on (depends on whether comment, free text, and mc_fc_vars)
    merge_cols = ["raw_value"]
    if free_text:
        merge_cols.append("raw_value_fc")
    if comment:
        merge_cols.append("comment")
    if mc_fc_vars and comment:
        merge_cols.append("comment_fc")

    # Create a rename dict to rename the relevant cols in the main dataset
    if not mc_fc_vars:
        rename_dict = {var_name: "raw_value"}
        if comment:
            rename_dict[f"{var_name}_co"] = "comment"
        if free_text:
            rename_dict[f"{var_name}_fc"] = "raw_value_fc"
    if mc_fc_vars:
        rename_dict = {f"{var_name}_mc": "raw_value", 
                       f"{var_name}_fc": "raw_value_fc"}
        if comment:
            rename_dict[f"{var_name}_mc_co"] = "comment"
            rename_dict[f"{var_name}_fc_co"] = "comment_fc"

    # Create re-renaming dict to rename back
    rename_back_dict = {"raw_value": var_name,
                        "comment": f"{var_name}_co",
                        "raw_value_fc": f"{var_name}_fc",
                        "comment_fc": f"{var_name}_fc_co"}
    
    # Helper function to enforce correct dtypes
    def enforce_dtypes(_df):

        out = _df.copy()

        # comment and free text cols should be string
        string_cols = [c for c in ["comment", "comment_fc", "raw_value_fc"] if c in out.columns]
        for c in string_cols:
            out[c] = out[c].astype("string").str.strip()
            out.loc[out[c] == "", c] = pd.NA
        
        # main raw value follows selected dtype
        if "raw_value" in out.columns:
            if dtype == "numeric":
                out["raw_value"] = pd.to_numeric(out["raw_value"], errors="coerce")
            elif dtype == "date":
                out["raw_value"] = pd.to_datetime(out["raw_value"], errors="coerce").dt.normalize()
            else:  # string
                out["raw_value"] = out["raw_value"].astype("string").str.strip()
                out.loc[out["raw_value"] == "", "raw_value"] = pd.NA

        return out

    # Quick check to make sure all the columns to merge on are in the main dataset
    required_source_cols = list(rename_dict.keys())
    missing_cols = [c for c in required_source_cols if c not in df.columns]
    if missing_cols:
        raise ValueError(
            f"Missing required column(s) in main dataset: {missing_cols}. "
            f"Expected based on settings: {required_source_cols}"
        )
    
    # Rename relevant cols in main dataset to match cleaning sheet for merging
    df = df.rename(columns=rename_dict).copy()
    df = enforce_dtypes(df)

    # Check that relevant cols are there after renaming
    missing_merge_cols = [c for c in merge_cols if c not in df.columns]
    if missing_merge_cols:
        raise ValueError(
            f"Missing merge column(s) after renaming: {missing_merge_cols}. "
            f"Available columns: {list(df.columns)}"
        )

    # Keep only relevant cols for merging and drop duplicates to get unique combinations, drop any with NA in all merge cols
    df_subset = df[merge_cols].drop_duplicates().reset_index(drop=True).copy()
    df_subset = df_subset.dropna(subset=merge_cols, how="all") # Drop any rows where variable, comment, and free text are all NA
    
    # Step 2: Merge with the cleaning worksheet to check if combination has already been added

    # Load the cleaning sheet
    wb = load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} does not exist. Please create it first.")
    ws = wb[sheet_name]
    cleaning_df = pd.DataFrame(ws.values)
    cleaning_df.columns = cleaning_df.iloc[0]  # Set the first row as variable names
    cleaning_df = cleaning_df[1:].copy()  # Remove the header row from the data
    cleaning_df = enforce_dtypes(cleaning_df)

    # Merge the unique values dataframe with the cleaning sheet dataframe
    merged_df = df_subset.merge(cleaning_df, on=merge_cols, how="left", indicator=True)

    # If "_merge" == "left_only", set cleaned_value and status to "" and "Unchecked" respectively
    merged_df.loc[merged_df["_merge"] == "left_only", "cleaned_value"] = ""
    merged_df.loc[merged_df["_merge"] == "left_only", "status"] = "Unchecked"

    # Step 3: If not already in cleaning sheet (i.e. _merge = "left_only"), add to the 
    # bottom of the cleaning sheet with status "Unchecked" for review

    new_rows = merged_df[merged_df["_merge"] == "left_only"].copy()

    # Append ONLY new rows (no header)
    if not new_rows.empty:
        # keep only sheet columns, in sheet order
        sheet_headers = [c.value for c in ws[1]]
        rows_to_append = new_rows.drop(columns=["_merge"], errors="ignore").copy()
        for col in sheet_headers:
            if col not in rows_to_append.columns:
                rows_to_append[col] = ""
        rows_to_append = rows_to_append[sheet_headers]

        # Ensure no pd.NA to avoid issues with excel
        rows_to_append = rows_to_append.fillna("")

        for r in dataframe_to_rows(rows_to_append, index=False, header=False):
            ws.append(r)

    wb.save(file_name)
    wb.close()

    # Step 4: Create clean and status variables in main dataset from cleaning sheet

    # Include rows just appended in Step 3 (so new combinations are available immediately)
    lookup_parts = [cleaning_df[merge_cols + ["cleaned_value", "status"]]]
    if not new_rows.empty:
        lookup_parts.append(
            new_rows.drop(columns=["_merge"], errors="ignore")[merge_cols + ["cleaned_value", "status"]]
        )

    lookup_df = pd.concat(lookup_parts, ignore_index=True)
    lookup_df = lookup_df.drop_duplicates(subset=merge_cols, keep="first")

    # Merge lookup onto df using merge keys
    df = df.merge(lookup_df, on=merge_cols, how="left")

    # Rename back to variable-specific names
    rename_out = {
        "cleaned_value": f"{var_name}_clean",
        "status": f"{var_name}_status",
    }
    for generic_col, original_col in rename_back_dict.items():
        if generic_col in df.columns:
            rename_out[generic_col] = original_col

    df = df.rename(columns=rename_out)
    
    # Step 5: If specified, produce short report of cleaning progress (e.g. no unique values, % cleaned, % pending, % exclude, etc.)
    if report:
        total_unique = len(merged_df)
        num_cleaned = len(merged_df[merged_df["status"] == "Cleaned"])
        num_exclude = len(merged_df[merged_df["status"] == "Exclude"])
        num_pending = len(merged_df[merged_df["status"] == "Pending"])
        num_unchecked = len(merged_df[merged_df["status"] == "Unchecked"])

        print(f"Cleaning progress for {var_name}:")
        print(f"Total unique value combinations: {total_unique}")
        print(f"Cleaned combinations: {num_cleaned}")
        print(f"Pending combinations: {num_pending}")
        print(f"Excluded combinations: {num_exclude}")
        print(f"Unchecked combinations: {num_unchecked}")
    return df