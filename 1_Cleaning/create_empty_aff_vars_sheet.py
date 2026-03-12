# Function to create empty excel sheets for cleaning unique values
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os

def create_empty_aff_vars_sheet(file_name, sheet_name, id_cols=None):
    
    # Create an empty dataframe with the specified columns
    if id_cols is None:
        id_cols = ["labgroupid"]

    # Define columns for the cleaning workbook
    columns = id_cols + [
        "trigger_var",
        "affected_var",
        "original_value",
        "cleaned_value",
        "check_status",
        "value_changed"
    ]
    
    # Check if the workbook already exists - if not, create it and create READ_ME sheet
    if not os.path.exists(file_name):

        wb = Workbook()

         # Rename default sheet to README
        ws_readme = wb.active
        ws_readme.title = "README"

        ws_readme["A1"] = "Cleaning Workbook"
        ws_readme["A1"].font = Font(bold=True)

        ws_readme["A3"] = "This workbook is for tracking the cleaning of affected vars"
        ws_readme["A4"] = "Fill in the cleaned value, check status ('Unchecked', 'Checked', 'Pending'), and whether the value was changed (Y/N)."

        ws_readme["A6"] = "Workflow:"
        ws_readme["A7"] = "- Scroll to the first row in which status is 'Unchecked'."
        ws_readme["A8"] = "- If can clean value, update cleaned_value and set status to 'Cleaned'."
        ws_readme["A9"] = "- If cannot currently clean value, set status to 'Pending'."

        ws_readme.column_dimensions["A"].width = 100

        wb.save(file_name)
        wb.close()

    # Load workbook
    wb = load_workbook(file_name)
    
    # Check if the sheet already exists - if not, create it
    if sheet_name not in wb.sheetnames:

        ws = wb.create_sheet(title=sheet_name)

        # Write headers and make them bold
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

         # Freeze panes
        ws.freeze_panes = "A2"

        # Make the columns sufficiently wide for headers (50 for original_value and cleaned_value, 20 for others)
        col_widths = {
            "original_value": 50,
            "cleaned_value": 50
        }

        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            width = col_widths.get(col_name, 20)  # Default width is 20 for other columns
            ws.column_dimensions[col_letter].width = width

        # Save the workbook
        wb.save(file_name)
        wb.close()

    else:
        print(f"Sheet {sheet_name} already exists. No changes made.")
        wb.close()