import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def _row_key(row, id_cols, sep="||"):
    parts = [str(row[c]).strip() for c in id_cols]
    return sep.join(parts)


def _ensure_split_sheet(file_name, sheet_name, headers):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "README"
        ws["A1"] = "Split Types Cleaning Workbook"
        ws["A3"] = "Workflow:"
        ws["A4"] = "1) Run notebook to export rows marked with status 'Split'."
        ws["A5"] = "2) Duplicate the 'Original' row(s), set split_type='Split', edit id/clean columns."
        ws["A6"] = "3) Re-run notebook to pull split rows back into the dataset."
        wb.save(file_name)
        wb.close()

    wb = load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, col_name in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
    else:
        ws = wb[sheet_name]

    # Keep workbook layout consistent and easy to read.
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(file_name)
    wb.close()


def clean_split_types(
    df,
    file_name,
    sheet_name,
    id_cols,
    clean_cols=None,
    split_status="Split",
    auto_assign_type_no=True,
    split_type_start=10,
):
    """
    Export rows where any *_status == split_status to Excel and pull back user-defined split rows.

    Sheet format:
    - split_type: "Original" or "Split"
    - original_key: immutable key copied from exported original rows
    - id columns (editable for Split rows)
    - clean columns (editable for Split rows)
    """
    out = df.copy()

    missing_ids = [c for c in id_cols if c not in out.columns]
    if missing_ids:
        raise ValueError(f"Missing identifier column(s): {missing_ids}")

    if out[id_cols].isna().any(axis=None):
        raise ValueError("Identifier columns contain NA; cannot safely split/merge rows.")

    if out.duplicated(subset=id_cols).any():
        raise ValueError("Identifier columns must uniquely identify rows before splitting.")

    status_cols = [c for c in out.columns if c.endswith("_status")]
    if not status_cols:
        print("No *_status columns found. Nothing to split.")
        return out

    if clean_cols is None:
        clean_cols = [c for c in out.columns if c.endswith("_clean")]

    split_mask = out[status_cols].eq(split_status).any(axis=1)
    candidates = out.loc[split_mask].copy()
    if candidates.empty:
        print(f"No rows found with any *_status == '{split_status}'.")
        return out

    candidates["original_key"] = candidates.apply(lambda r: _row_key(r, id_cols), axis=1)

    headers = ["split_type", "original_key"] + id_cols + clean_cols
    _ensure_split_sheet(file_name=file_name, sheet_name=sheet_name, headers=headers)

    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    rows = list(ws.values)
    sheet_cols = list(rows[0]) if rows else headers
    existing = pd.DataFrame(rows[1:], columns=sheet_cols) if len(rows) > 1 else pd.DataFrame(columns=sheet_cols)

    required = ["split_type", "original_key"] + id_cols + clean_cols
    missing_sheet_cols = [c for c in required if c not in existing.columns]
    if missing_sheet_cols:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' missing required columns: {missing_sheet_cols}")

    export_original = candidates[["original_key"] + id_cols + clean_cols].copy()
    export_original.insert(0, "split_type", "Original")

    existing_original_keys = set(
        existing.loc[
            existing["split_type"].astype(str).str.strip().str.lower() == "original",
            "original_key"
        ].astype(str)
    )
    to_append = export_original[~export_original["original_key"].astype(str).isin(existing_original_keys)].copy()

    if not to_append.empty:
        rows_to_write = to_append.copy()
        for c in sheet_cols:
            if c not in rows_to_write.columns:
                rows_to_write[c] = ""
        rows_to_write = rows_to_write[sheet_cols]
        rows_to_write = rows_to_write.where(rows_to_write.notna(), None)

        next_row = ws.max_row + 1
        for _, row in rows_to_write.iterrows():
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=next_row, column=col_idx, value=val)
            next_row += 1
        wb.save(file_name)
        print(f"Appended {len(to_append)} new Original row(s) to '{sheet_name}'.")

        existing = pd.concat([existing, to_append[sheet_cols]], ignore_index=True)
    else:
        print("No new Original rows to append.")

    wb.close()

    split_rows = existing[existing["split_type"].astype(str).str.strip().str.lower() == "split"].copy()
    split_rows = split_rows[split_rows["original_key"].notna()].copy()
    split_rows["original_key"] = split_rows["original_key"].astype(str).str.strip()

    # If requested, auto-fill missing type_no values for split rows.
    # Numbering is per original_key and starts at 1 in workbook row order.
    if auto_assign_type_no and "type_no" in id_cols and "type_no" in split_rows.columns:
        split_rows["type_no"] = split_rows["type_no"].astype("string").str.strip()
        split_rows.loc[split_rows["type_no"] == "", "type_no"] = pd.NA

        missing_mask = split_rows["type_no"].isna()
        if missing_mask.any():
            seq = (
                split_rows.loc[missing_mask]
                .groupby("original_key", sort=False)
                .cumcount()
                + int(split_type_start)
            )
            split_rows.loc[missing_mask, "type_no"] = seq.astype("Int64").astype("string")

        split_rows["type_no"] = pd.to_numeric(split_rows["type_no"], errors="coerce")

    if split_rows.empty:
        print("No Split rows found in workbook yet. Export completed.")
        return out

    base_lookup = candidates.set_index("original_key", drop=False)
    valid_split_rows = split_rows[split_rows["original_key"].isin(base_lookup.index)].copy()

    if valid_split_rows.empty:
        print("Split rows found, but none match exported original_key values.")
        return out

    if valid_split_rows.duplicated(subset=id_cols).any():
        raise ValueError(
            "Split rows produce duplicate identifiers. Ensure each split row has a unique "
            f"combination of id_cols: {id_cols}"
        )

    new_rows = []
    for _, split_row in valid_split_rows.iterrows():
        base = base_lookup.loc[split_row["original_key"]].copy()
        for col in id_cols + clean_cols:
            if col in split_row.index and pd.notna(split_row[col]):
                base[col] = split_row[col]
        new_rows.append(base)

    new_rows_df = pd.DataFrame(new_rows)
    keys_to_replace = set(valid_split_rows["original_key"].tolist())
    out_keys = out.apply(lambda r: _row_key(r, id_cols), axis=1)
    out_remaining = out.loc[~out_keys.isin(keys_to_replace)].copy()

    combined = pd.concat([out_remaining, new_rows_df[out.columns]], ignore_index=True)
    print(
        f"Replaced {len(keys_to_replace)} original row(s) with {len(new_rows_df)} split row(s)."
    )

    return combined


def reassign_type_no(
    df,
    group_cols=("labgroupid", "equipment", "survey"),
    type_col="type_no",
    start=1,
):
    """
    Reassign type_no sequentially within each (labgroupid, equipment, survey) group.
    Useful as a final step after exclusions/splits.
    """
    out = df.copy()

    required_cols = list(group_cols) + [type_col]
    missing_cols = [c for c in required_cols if c not in out.columns]
    if missing_cols:
        raise ValueError(f"Missing required column(s): {missing_cols}")

    # Keep stable ordering by current type_no where possible.
    out["_type_sort"] = pd.to_numeric(out[type_col], errors="coerce")
    out = out.sort_values(list(group_cols) + ["_type_sort", type_col], kind="stable").copy()

    out[type_col] = (
        out.groupby(list(group_cols), sort=False).cumcount() + int(start)
    )

    out = out.drop(columns=["_type_sort"])
    return out
