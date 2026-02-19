# Function to create empty excel sheets for cleaning unique values
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os

def create_empty_cleaning_sheet(file_name, sheet_name,
                                comment=False, 
                                free_text=False, 
                                mc_fc_vars=False):
    
    # Create an empty dataframe with the specified columns
    columns = ["raw_value"]
    if free_text:
        columns.append("raw_value_fc")
    if comment:
        columns.append("comment")
    if mc_fc_vars and comment:
        columns.append("comment_fc")
    
    # Add checking columns
    columns.extend([
        "cleaned_value",
        "status",
        "needs_review",
        "affects_vars",
        "dependent_on_vars",
        "combine_group",
        "decision_rule",
        "notes",
    ])

    # Check if the workbook already exists - if not, create it and create READ_ME sheet
    if not os.path.exists(file_name):
        wb = Workbook()

         # Rename default sheet to README
        ws_readme = wb.active
        ws_readme.title = "README"

        ws_readme["A1"] = "Cleaning Workbook"
        ws_readme["A1"].font = Font(bold=True)

        ws_readme["A3"] = "Instructions:"
        ws_readme["A4"] = "- Each sheet contains unique raw values and comment combinations."
        ws_readme["A5"] = "- Update status and review fields as needed."
        ws_readme["A6"] = "- Do not modify column headers."
        ws_readme["A7"] = "- Do not edit raw values or comments."

        ws_readme["A9"] = "Workflow:"
        ws_readme["A10"] = "- Scroll to the first row in which status is 'Unchecked'."
        ws_readme["A11"] = "- If can clean value, update cleaned_value and set status to 'Cleaned'."
        ws_readme["A12"] = "- If cannot currently clean value, set status to 'Pending'."
        ws_readme["A13"] = "- If should exclude value, leave cleaned_value blank and set status to 'Exclude'."
        ws_readme["A14"] = "- If cannot clean as dependent on other variables, put 'Y' in the 'dependent_on_vars' column."
        ws_readme["A15"] = "- If refers to other obs, put the group name in the 'combine_group' column."
        ws_readme["A16"] = "- Add any decision rules or notes in the respective columns."
        ws_readme["A17"] = "- If affects other variables, list them in the 'affects_vars' column, separated by commas."

        ws_readme.column_dimensions["A"].width = 100

        wb.save(file_name)
        wb.close()

    # Load workbook
    wb = load_workbook(file_name)
    
    # Check if the sheet already exists - if not, create it
    if sheet_name not in wb.sheetnames:

        ws = wb.create_sheet(title=sheet_name)

        # Write headers and make them bold
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

         # Freeze panes
        ws.freeze_panes = "A2"

        # Make the columns sufficiently wide for headers (50 for raw values and comments, 20 for others)
        col_widths = {
            "raw_value": 50,
            "raw_value_fc": 50,
            "comment": 50,
            "comment_fc": 50,
        }

        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            width = col_widths.get(col_name, 20)  # Default width is 20 for other columns
            ws.column_dimensions[col_letter].width = width

        # Save the workbook
        wb.save(file_name)
        wb.close()

    else:
        print(f"Sheet {sheet_name} already exists. No changes made.")
        wb.close()